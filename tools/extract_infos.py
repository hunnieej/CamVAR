# coding=utf-8
import os
import json
import sys

import xlwt
import openpyxl
from glob import glob
from tqdm import tqdm


class ExtractInfos(object):
    def __init__(self):
        self.remove_character = [
            # '\u0005',
            # '\u0002',
            # '\u0016',
            # '\u0001',
            # '\u000b'
            '\b',
        ]
        self.prompt_none = '11==22==33'

    # def find_illegal_char(self, path):
    #     with open(path, 'r') as file:
    #         lines = file.readlines()
    #         progress_bar = tqdm(range(len(lines)), total=len(lines))
    #         progress_bar.set_description("Processing")
    #
    #         for line in lines:
    #             line = line.strip('\n')
    #             if '\\u' in line:
    #                 try:
    #                     line = line.encode('utf-8').decode("unicode_escape")
    #                 except Exception as e:
    #                     print(e)
    #                     print(line)
    #             progress_bar.update(1)

    def extract_prompt(self, json_dir, save_dir):
        """
        从 json 文件中抽取出 prompt 信息，并将其保存到 txt 文件中
        """
        os.makedirs(save_dir, exist_ok=True)
        base_dir = os.path.basename(json_dir)
        save_file = os.path.join(save_dir, base_dir + '.txt')
        json_paths = glob(os.path.join(json_dir, '*.json'))
        with open(save_file, 'w', encoding='UTF-8') as out_file:
            for i, json_path in enumerate(json_paths):
                file_name = os.path.basename(json_path)
                print('Current handle file {} {}'.format(file_name, i))

                with open(json_path, 'r') as file:
                    data = json.load(file)
                    for key in data.keys():
                        try:
                            language = data.get(key).get('LANGUAGE')
                            if language != 'zh':
                                continue 
                            prompt = data.get(key).get('prompt')
                            if prompt is None:
                                prompt = self.prompt_none
                            else:
                                # if '\\u' in prompt:
                                #     prompt = prompt.encode('utf-8').decode("unicode_escape")

                                for rm_c in self.remove_character:
                                    prompt = prompt.replace(rm_c, ' ')
                        except Exception as e:
                            prompt = self.prompt_none  # 将有问题的 prompt 统一转换为字符串 11==22==33

                        res = file_name + '\t' + key + '\t' + prompt + '\n'
                        out_file.write(res)

    def write_xls(self, input_file, output_dir):
        """
        将文本中的 prompt 写入到 xls 文件中
        """
        os.makedirs(output_dir, exist_ok=True)
        basename = os.path.basename(input_file)
        short_name = os.path.splitext(basename)[0]
        output_file = os.path.join(output_dir, short_name + '.xls')

        # 创建一个 workbook 设置编码
        workbook = xlwt.Workbook(encoding='utf-8')
        # 创建一个 worksheet
        worksheet = workbook.add_sheet('Sheet1')

        with open(input_file, 'r') as file:
            for i, line in enumerate(file.readlines()):
                line = line.strip('\n')
                items = line.split('\t')
                prompt = items[-1]
                print(i, prompt)
                worksheet.write(i, 0, prompt)

        workbook.save(output_file)

    def write_xlsx_(self, prompts, output_file):
        # 创建一个Workbook对象，相当于创建了一个Excel文件
        workbook = openpyxl.Workbook()
        # workbook = openpyxl.Workbook(encoding='UTF-8')

        # 获取当前活跃的worksheet,默认就是第一个worksheet
        worksheet = workbook.active
        worksheet.title = "Sheet1"

        # worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
        # # worksheet2 = workbook.create_sheet(0)  #插入在工作簿的第一个位置
        # worksheet2.title = "New Title"

        for i, prompt in enumerate(prompts):
            try:
                prompt = prompt.encode('utf-8').decode("unicode_escape")
                worksheet.cell(i + 1, 1, prompt)
            except Exception as e:
                worksheet.cell(i + 1, 1, self.prompt_none)  # 如果转换 prompt 或写入 prompt 报错，则写入prompt 的空表示符

        workbook.save(filename=output_file)

    def write_xlsx(self, input_file, output_dir):
        """
        将文本中的 prompt 写入到 xlsx 文件中
        """
        os.makedirs(output_dir, exist_ok=True)
        basename = os.path.basename(input_file)
        short_name = os.path.splitext(basename)[0]
        output_file = os.path.join(output_dir, short_name + '.xlsx')

        prompts = []
        with open(input_file, 'r') as file:
            for i, line in enumerate(file.readlines()):
                print(i)
                line = line.strip('\n')
                items = line.split('\t')
                prompt = items[-1]
                prompts.append(prompt)
        self.write_xlsx_(prompts, output_file)

    def write_multi_xlsx(self, input_file, output_dir, num=100000):
        """
        将文本中的 prompt 写入到 xlsx 文件中
        """

        basename = os.path.basename(input_file)
        short_name = os.path.splitext(basename)[0]
        output_dir = os.path.join(output_dir, short_name)
        os.makedirs(output_dir, exist_ok=True)

        prompts = []
        with open(input_file, 'r') as file:
            for line in file.readlines():
                line = line.strip('\n')
                items = line.split('\t')
                prompt = items[-1]
                prompts.append(prompt)

        total_num = len(prompts)
        a = total_num // num
        b = total_num % num

        if b > 0:
            a += 1
        progress_bar = tqdm(range(a), total=a)
        progress_bar.set_description("Writing xlsx")
        for i in range(a):
            start = i * num
            end = (i + 1) * num
            end = min(end, total_num)

            prompts_ = prompts[start:end]
            output_file = os.path.join(output_dir, '{}.xlsx'.format(i))
            self.write_xlsx_(prompts_, output_file)

            progress_bar.update(1)

    def read_xlsx(self, file):
        """
        从 xlsx 文件中读取数据
        """
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook["工作表1"]
        # rows = worksheet.max_row
        # columns = worksheet.max_column

        res = []
        i = 1
        for row in worksheet.rows:
            for cell in row:
                if cell.value == '' or cell.value is None:
                    print(file, i)
                    res.append("")
                else:
                    res.append(cell.value)
            i += 1
        return res

    def mapping(self, txt_file, excel_path, output_dir):
        """
        将 excel 文件中的翻译结果映射到 txt 文件中
        """
        os.makedirs(output_dir, exist_ok=True)
        basename = os.path.basename(txt_file)
        output_file = os.path.join(output_dir, basename)
        if os.path.isfile(excel_path):
            zhs = self.read_xlsx(excel_path)
        elif os.path.isdir(excel_path):
            zhs = []
            paths = glob(os.path.join(excel_path, '*.xlsx'))
            for i in range(len(paths)):
                print('Read excel file {}'.format(i))
                path = os.path.join(os.path.dirname(paths[i]), '{}.xlsx'.format(i))
                zhs_ = self.read_xlsx(path)
                zhs.extend(zhs_)
        else:
            raise ValueError('The excel path {} is not support.'.format(excel_path))

        with open(txt_file, 'r') as file, open(output_file, 'w') as out_file:
            lines = file.readlines()
            assert len(lines) == len(zhs)

            for i, line in enumerate(lines):
                line = line.strip('\n')
                line += '\t' + zhs[i] + '\n'
                out_file.write(line)

    def convert(self, txt_file, json_dir, save_dir, field_name):
        """
        将 txt 文件中的翻译结果映射到 json 文件中
        """
        base_dir = os.path.basename(json_dir)
        save_dir = os.path.join(save_dir, base_dir)
        os.makedirs(save_dir, exist_ok=True)

        translation_infos = {}
        with open(txt_file, 'r') as file:
            for line in file.readlines():
                line = line.strip('\n')
                items = line.split('\t')
                assert len(items) == 4
                json_file = items[0]
                image_name = items[1]
                translation_info = items[-1]
                if translation_info == self.prompt_none:
                    translation_info = None

                if json_file in translation_infos:
                    translation_infos[json_file][image_name] = translation_info
                else:
                    translation_infos[json_file] = {}
                    translation_infos[json_file][image_name] = translation_info

        json_paths = glob(os.path.join(json_dir, '*.json'))

        for json_path in json_paths:
            file_name = os.path.basename(json_path)
            print('Current handle file {}'.format(file_name))

            res = {}
            with open(json_path, 'r') as file:
                data = json.load(file)
                for image_name in data.keys():
                    # data[image_name][field_name] = translation_infos[file_name][image_name]

                    if image_name not in res:
                        res[image_name] = {}
                    res[image_name]['prompt'] = data[image_name]['prompt']
                    res[image_name][field_name] = translation_infos[file_name][image_name]

            new_file_name = '-'.join(file_name.split('-')[:2]) + '-prompt.json'
            out_path = os.path.join(save_dir, new_file_name)

            with open(out_path, 'w') as file:
                file.write(json.dumps(res, indent=4, ensure_ascii=False))
